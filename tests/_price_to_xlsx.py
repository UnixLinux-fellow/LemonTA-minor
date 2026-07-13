#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Convert tests/price.json to tests/price.xlsx (multi-sheet)."""
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(HERE, "price.json")
DST = os.path.join(HERE, "price.xlsx")

CATEGORY_CN = {
    "hinge": "铰链",
    "hangingRail": "挂衣杆",
    "pushLatch": "反弹器",
    "supportArm": "支撑杆",
    "lightingSystem": "灯光系统",
    "minifix": "三合一连接件",
    "slide": "滑轨",
    "plinthSystem": "踢脚系统",
    "doorBumper": "门碰",
    "liquidNails": "免钉胶",
    "selfTappingScrew": "自攻螺丝",
    "accessPanelHandle": "检修口拉手",
    "nylonPre-InsertedNut ": "尼龙预埋螺母",
    "dustStrip": "防尘胶条",
    "hingeBaseCap": "铰链底盖",
    "cabinetBodyConnector": "柜体连接件",
}

HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="4F81BD")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER


def style_body(ws, nrows, ncols, left_cols=None):
    left_cols = left_cols or []
    for r in range(2, nrows + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = LEFT if c in left_cols else CENTER
            cell.border = BORDER


def autosize(ws, widths):
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = w


def write_simple(ws, mapping, unit):
    ws.append(["名称", f"单价（{unit}）"])
    for name, price in mapping.items():
        ws.append([name, price])
    style_header(ws, 2)
    style_body(ws, ws.max_row, 2, left_cols=[1])
    autosize(ws, [32, 18])
    ws.freeze_panes = "A2"


def write_hardware(ws, hardware):
    ws.append(["类别", "英文键", "物料名称", "单价（元）"])
    for eng_cat, items in hardware.items():
        cn_cat = CATEGORY_CN.get(eng_cat, eng_cat)
        for item, price in items.items():
            ws.append([cn_cat, eng_cat, item, price])
    style_header(ws, 4)
    style_body(ws, ws.max_row, 4, left_cols=[1, 2, 3])
    autosize(ws, [16, 22, 60, 14])
    ws.freeze_panes = "A2"


def main():
    with open(SRC, "r", encoding="utf-8") as f:
        data = json.load(f)
    pl = data["priceList"]

    wb = Workbook()
    wb.remove(wb.active)

    write_simple(wb.create_sheet("板材单价"), pl["panelPrice"], "元/㎡")
    write_simple(wb.create_sheet("门板材质"), pl["doorMaterial"], "元/㎡")
    write_simple(wb.create_sheet("门板工艺"), pl["doorCraft"], "元/㎡")

    for brand, items in pl["hardware"].items():
        write_hardware(wb.create_sheet(f"五金-{brand}"), items)

    wb.save(DST)
    print(f"OK -> {DST}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
