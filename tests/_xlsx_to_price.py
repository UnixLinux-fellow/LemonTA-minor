#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Parse tests/price.xlsx back into tests/price_list.json."""
import json
import os
from collections import OrderedDict
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(HERE, "price.xlsx")
DST = os.path.join(HERE, "price_list.json")


def read_simple(ws):
    """Two-column sheet: 名称 / 单价. Returns OrderedDict."""
    out = OrderedDict()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        name, price = row[0], row[1]
        if name is None or price is None:
            continue
        out[str(name).strip()] = float(price)
    return out


def read_hardware(ws):
    """Four-column sheet: 类别 / 英文键 / 物料名称 / 单价. Returns OrderedDict keyed by 英文键."""
    out = OrderedDict()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        _cn, eng, item, price = row[0], row[1], row[2], row[3]
        if eng is None or item is None or price is None:
            continue
        eng_key = str(eng)  # 保留原始尾随空格情况
        # 若 sheet 里删掉了尾随空格,回原 key（nylonPre-InsertedNut 带尾空格）;这里按 sheet 原样输出
        bucket = out.setdefault(eng_key, OrderedDict())
        bucket[str(item).strip()] = float(price)
    return out


def normalize_number(v):
    """Keep integers as-is (0), floats trimmed of trailing .0 stays float in JSON."""
    return v


def main():
    wb = load_workbook(SRC, data_only=True)
    names = wb.sheetnames

    def find(candidates):
        for c in candidates:
            if c in names:
                return wb[c]
        raise KeyError(f"未找到 sheet: {candidates}")

    panel_ws = find(["板材单价", "板材"])
    door_mat_ws = find(["门板材质"])
    door_craft_ws = find(["门板工艺"])

    hardware = OrderedDict()
    for name in names:
        if name.startswith("五金-"):
            brand = name.split("-", 1)[1]
            hardware[brand] = read_hardware(wb[name])

    price_list = OrderedDict()
    price_list["panelPrice"] = read_simple(panel_ws)
    price_list["doorMaterial"] = read_simple(door_mat_ws)
    price_list["doorCraft"] = read_simple(door_craft_ws)
    price_list["hardware"] = hardware

    data = OrderedDict()
    data["priceList"] = price_list

    with open(DST, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"OK -> {DST}")

    # 打印统计
    print(f"  panelPrice   : {len(price_list['panelPrice'])} 项")
    print(f"  doorMaterial : {len(price_list['doorMaterial'])} 项")
    print(f"  doorCraft    : {len(price_list['doorCraft'])} 项")
    for brand, items in hardware.items():
        total = sum(len(v) for v in items.values())
        print(f"  hardware[{brand}] : {len(items)} 类 / {total} 项")


if __name__ == "__main__":
    main()
